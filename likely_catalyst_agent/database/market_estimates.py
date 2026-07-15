"""Parse Bloomberg EEG / ERN workbooks into ontology-ready row dicts."""

from __future__ import annotations

import logging
import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from src.pipeline import _derive_period

log = logging.getLogger(__name__)

_DEFAULT_TICKER = "AAPL"
_EEG_SHEET = "Sheet1"
_ERN_SHEET = "Worksheet"
_EEG_HEADER_ROWS = 2
_SOURCE_BQL = "BLOOMBERG_BQL"
_SOURCE_BDH = "BLOOMBERG_BDH"
_SOURCE_ERN = "BLOOMBERG_ERN"

# Fallback when BQL formula text is missing (column letter -> FPR year, FPT).
_EEG_COLUMN_SPEC: dict[str, dict[str, Any]] = {
    "A": {"fpr": 2026, "fpt": "A"},
    "D": {"fpr": 2027, "fpt": "A"},
}

_EEG_ESTIMATE_BLOCKS: tuple[tuple[str, str], ...] = (("A", "B"), ("D", "E"))
_FPR_RE = re.compile(r"FPR=(\d{4})Y")
_FPT_RE = re.compile(r"FPT=([AQ])")
_AS_OF_RANGE_RE = re.compile(
    r"AS_OF_DATE=RANGE\(\s*(\d{4}-\d{2}-\d{2})\s*,\s*(\d{4}-\d{2}-\d{2})\s*\)"
)


def _formula_text(cell_value: Any) -> str | None:
    if cell_value is None:
        return None
    if hasattr(cell_value, "text"):
        return str(cell_value.text)
    if isinstance(cell_value, str):
        return cell_value
    return str(cell_value)


def _to_date(value: Any) -> date | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    return pd.to_datetime(value).date()


def _parse_pull_params(formula_text: str | None, col_letter: str) -> dict[str, Any]:
    spec = dict(_EEG_COLUMN_SPEC.get(col_letter, {}))
    if formula_text:
        m_fpr = _FPR_RE.search(formula_text)
        m_fpt = _FPT_RE.search(formula_text)
        if m_fpr:
            spec["FPR"] = m_fpr.group(1)
        if m_fpt:
            spec["FPT"] = m_fpt.group(1)
        m_rng = _AS_OF_RANGE_RE.search(formula_text)
        if m_rng:
            spec["AS_OF_RANGE"] = [m_rng.group(1), m_rng.group(2)]
        if "F1396" in formula_text:
            spec["source_field"] = "F1396"
        if "ACT_EST_MAPPING" in formula_text:
            m = re.search(r"ACT_EST_MAPPING='([^']*)'", formula_text)
            if m:
                spec["ACT_EST_MAPPING"] = m.group(1)
    return spec


def _target_from_params(pull_params: dict[str, Any]) -> tuple[str, str]:
    fpr = pull_params.get("FPR") or pull_params.get("fpr")
    fpt = pull_params.get("FPT") or pull_params.get("fpt") or "A"
    if fpr is None:
        raise ValueError("missing FPR in pull_params")
    year = int(fpr) if not isinstance(fpr, int) else fpr
    if str(fpt).upper() == "A":
        return _derive_period(year, None), "A"
    return _derive_period(year, 1), str(fpt).upper()


def parse_eeg(path: str | Path, *, ticker: str = _DEFAULT_TICKER) -> list[dict[str, Any]]:
    """Parse consensus estimate time series from an EEG workbook."""
    path = Path(path)
    wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws_f = wb_formula[_EEG_SHEET]
    ws_v = wb_values[_EEG_SHEET]
    rows_out: list[dict[str, Any]] = []

    try:
        for date_col, value_col in _EEG_ESTIMATE_BLOCKS:
            formula = _formula_text(ws_f[f"{date_col}1"].value)
            if not formula:
                log.warning(
                    "EEG %s: no formula in %s1; using _EEG_COLUMN_SPEC fallback",
                    path.name,
                    date_col,
                )
            pull_params = _parse_pull_params(formula, date_col)
            target_period, target_period_type = _target_from_params(pull_params)

            pairs: list[tuple[date, float]] = []
            for row_idx in range(_EEG_HEADER_ROWS + 1, ws_v.max_row + 1):
                d = _to_date(ws_v[f"{date_col}{row_idx}"].value)
                v = ws_v[f"{value_col}{row_idx}"].value
                if d is None or v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                pairs.append((d, float(v)))

            for as_of_date, value_mean in pairs:
                rows_out.append(
                    {
                        "ticker": ticker,
                        "metric": "EPS",
                        "target_period": target_period,
                        "target_period_type": target_period_type,
                        "as_of_date": as_of_date,
                        "value_mean": value_mean,
                        "value_high": None,
                        "value_low": None,
                        "value_median": None,
                        "n_estimates": None,
                        "value_stdev": None,
                        "currency": "USD",
                        "source": _SOURCE_BQL,
                        "source_field": pull_params.get("source_field", "F1396"),
                        "pull_params": pull_params,
                    }
                )
    finally:
        wb_formula.close()
        wb_values.close()

    return rows_out


def parse_eeg_price(path: str | Path, *, ticker: str = _DEFAULT_TICKER) -> list[dict[str, Any]]:
    """Parse daily close prices from cols G/H of an EEG workbook."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[_EEG_SHEET]
    rows_out: list[dict[str, Any]] = []
    try:
        for row_idx in range(1, ws.max_row + 1):
            price_date = _to_date(ws[f"G{row_idx}"].value)
            close_px = ws[f"H{row_idx}"].value
            if price_date is None or close_px is None:
                continue
            if isinstance(close_px, float) and pd.isna(close_px):
                continue
            rows_out.append(
                {
                    "ticker": ticker,
                    "price_date": price_date,
                    "close_px": float(close_px),
                    "currency": "USD",
                    "source": _SOURCE_BDH,
                    "source_field": "PR005",
                }
            )
    finally:
        wb.close()
    return rows_out


def _parse_per_label(per: str) -> tuple[int, int]:
    m = re.match(r"Q(\d)\s+(\d{2})\b", str(per).strip(), re.IGNORECASE)
    if not m:
        raise ValueError(f"cannot parse Per label: {per!r}")
    quarter = int(m.group(1))
    yy = int(m.group(2))
    year = 2000 + yy if yy < 100 else yy
    return quarter, year


def _parse_per_end_month_end(per_end: Any) -> date | None:
    if per_end is None or (isinstance(per_end, float) and pd.isna(per_end)):
        return None
    if isinstance(per_end, (datetime, date)):
        d = per_end.date() if isinstance(per_end, datetime) else per_end
        last = monthrange(d.year, d.month)[1]
        return date(d.year, d.month, last)
    text = str(per_end).strip()
    if "/" not in text:
        return _to_date(per_end)
    month_s, year_s = text.split("/", 1)
    month = int(month_s)
    yy = int(year_s)
    year = 2000 + yy if yy < 100 else yy
    last_day = monthrange(year, month)[1]
    return date(year, month, last_day)


def _parse_ann_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    return pd.to_datetime(value).date()


def _parse_pct(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        text = value.strip().replace("%", "")
        if not text:
            return None
        return float(text)
    return float(value)


def _parse_float(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return float(value)


def parse_ern(path: str | Path, *, ticker: str = _DEFAULT_TICKER) -> list[dict[str, Any]]:
    """Parse earnings surprise rows from an ERN workbook."""
    path = Path(path)
    df = pd.read_excel(path, sheet_name=_ERN_SHEET, header=0)
    mask = df["Per"].notna() & df["Per End"].notna()
    df = df.loc[mask].copy()
    rows_out: list[dict[str, Any]] = []

    for _, row in df.iterrows():
        fiscal_quarter, fiscal_year = _parse_per_label(row["Per"])
        fiscal_period = _derive_period(fiscal_year, fiscal_quarter)
        reported = row.get("Reported")
        is_reported = bool(pd.notna(reported))

        rows_out.append(
            {
                "ticker": ticker,
                "fiscal_period": fiscal_period,
                "fiscal_year": fiscal_year,
                "fiscal_quarter": fiscal_quarter,
                "period_end_date": _parse_per_end_month_end(row["Per End"]),
                "announcement_date": _parse_ann_date(row["Ann Date"]),
                "reported_eps": _parse_float(reported) if is_reported else None,
                "comparable_eps": _parse_float(row.get("Comp")),
                "estimate_eps": _parse_float(row.get("Estimate")),
                "surprise_pct": _parse_pct(row.get("%Surp")),
                "guidance_eps": _parse_float(row.get("Guidance")),
                "guidance_surprise_pct": _parse_pct(row.get("%Guid Surp")),
                "price_change_pct": _parse_pct(row.get("%Px Chg")),
                "eps_ttm": _parse_float(row.get("T12M")),
                "pe_ratio": _parse_float(row.get("P/E")),
                "is_reported": is_reported,
                "source": _SOURCE_ERN,
                "pull_params": None,
            }
        )
    return rows_out
